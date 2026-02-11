import json
import base64
from datetime import datetime
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, WebSocket, WebSocketDisconnect, Request
from sqlalchemy import desc, and_
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Alarm, User, AIBox
from app.schemas import AlarmCreate, AlarmResponse, AlarmUpdate
from app.auth import get_current_user
from app.config import settings
from app.services.bmapp import add_client, remove_client, broadcast_alarm, BmAppAlarmListener
from app.services.minio_storage import get_minio_storage

router = APIRouter(prefix="/alarms", tags=["alarms"])


def _add_presigned_urls(alarm: Alarm) -> dict:
    """Add presigned URLs to alarm response."""
    data = {
        "id": alarm.id,
        "bmapp_id": alarm.bmapp_id,
        "alarm_type": alarm.alarm_type,
        "alarm_name": alarm.alarm_name,
        "camera_id": alarm.camera_id,
        "camera_name": alarm.camera_name,
        "location": alarm.location,
        "confidence": alarm.confidence,
        "image_url": alarm.image_url,
        "video_url": alarm.video_url,
        "media_url": alarm.media_url,  # RTSP URL for video source
        "description": alarm.description,
        "status": alarm.status,
        "alarm_time": alarm.alarm_time,
        "created_at": alarm.created_at,
        "acknowledged_at": alarm.acknowledged_at,
        "acknowledged_by_id": alarm.acknowledged_by_id,
        "resolved_at": alarm.resolved_at,
        "resolved_by_id": alarm.resolved_by_id,
        "raw_data": alarm.raw_data,  # Include raw_data for bounding box info
        "minio_image_path": alarm.minio_image_path,
        "minio_labeled_image_path": alarm.minio_labeled_image_path,
        "minio_video_path": alarm.minio_video_path,
        "minio_synced_at": alarm.minio_synced_at,
        "minio_image_url": None,
        "minio_labeled_image_url": None,
        "minio_video_url": None,
    }

    storage = get_minio_storage()
    if storage.is_initialized:
        if alarm.minio_image_path:
            data["minio_image_url"] = storage.get_presigned_url(
                settings.minio_bucket_alarm_images,
                alarm.minio_image_path
            )
        if alarm.minio_labeled_image_path:
            data["minio_labeled_image_url"] = storage.get_presigned_url(
                settings.minio_bucket_alarm_images,
                alarm.minio_labeled_image_path
            )
        if alarm.minio_video_path:
            data["minio_video_url"] = storage.get_presigned_url(
                settings.minio_bucket_alarm_images,
                alarm.minio_video_path
            )

    return data


@router.get("/", response_model=List[AlarmResponse])
def get_alarms(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    alarm_type: Optional[str] = None,
    camera_id: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all alarms with optional filters"""
    query = db.query(Alarm)

    filters = []
    if alarm_type:
        filters.append(Alarm.alarm_type == alarm_type)
    if camera_id:
        filters.append(Alarm.camera_id == camera_id)
    if status:
        filters.append(Alarm.status == status)
    if start_date:
        filters.append(Alarm.alarm_time >= start_date)
    if end_date:
        filters.append(Alarm.alarm_time <= end_date)

    if filters:
        query = query.filter(and_(*filters))

    alarms = query.order_by(desc(Alarm.alarm_time)).offset(skip).limit(limit).all()
    return [_add_presigned_urls(a) for a in alarms]


@router.get("/stats")
def get_alarm_stats(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get alarm statistics"""
    query = db.query(Alarm)

    if start_date:
        query = query.filter(Alarm.alarm_time >= start_date)
    if end_date:
        query = query.filter(Alarm.alarm_time <= end_date)

    total = query.count()
    new_count = query.filter(Alarm.status == "new").count()
    acknowledged_count = query.filter(Alarm.status == "acknowledged").count()
    resolved_count = query.filter(Alarm.status == "resolved").count()

    # Get count by type
    from sqlalchemy import func
    type_stats = db.query(
        Alarm.alarm_type,
        func.count(Alarm.id).label("count")
    ).group_by(Alarm.alarm_type).all()

    return {
        "total": total,
        "new": new_count,
        "acknowledged": acknowledged_count,
        "resolved": resolved_count,
        "by_type": {t.alarm_type: t.count for t in type_stats}
    }


def _get_severity(alarm_type: str) -> str:
    """Get severity level for alarm type.

    Returns same severity for all types since BM-APP doesn't send severity info.
    """
    from app.alarm_types import get_alarm_severity
    return get_alarm_severity(alarm_type).capitalize()


@router.get("/export/excel")
async def export_excel(
    alarm_type: Optional[str] = None,
    camera_id: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export alarms to Excel - for Catatan Pelanggaran."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    query = db.query(Alarm)
    filters = []
    if alarm_type:
        filters.append(Alarm.alarm_type == alarm_type)
    if camera_id:
        filters.append(Alarm.camera_id == camera_id)
    if status:
        filters.append(Alarm.status == status)
    if start_date:
        filters.append(Alarm.alarm_time >= start_date)
    if end_date:
        filters.append(Alarm.alarm_time <= end_date)
    if filters:
        query = query.filter(and_(*filters))

    alarms = query.order_by(desc(Alarm.alarm_time)).limit(1000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catatan Pelanggaran"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["No", "Waktu", "Kamera", "Tipe", "Severity", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for idx, alarm in enumerate(alarms, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=alarm.alarm_time.strftime('%d %b %Y, %H:%M') if alarm.alarm_time else '').border = thin_border
        ws.cell(row=row, column=3, value=alarm.camera_name or 'Unknown').border = thin_border
        ws.cell(row=row, column=4, value=alarm.alarm_type or '').border = thin_border
        ws.cell(row=row, column=5, value=_get_severity(alarm.alarm_type or '')).border = thin_border
        ws.cell(row=row, column=6, value=alarm.status or '').border = thin_border

    column_widths = [6, 20, 25, 18, 12, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    filename = f"catatan_pelanggaran_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/export/excel-images")
async def export_excel_images(
    alarm_type: Optional[str] = None,
    camera_id: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export alarms to Excel with images - for Bukti Foto."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import httpx

    # Limit max to 100 to prevent timeout
    limit = min(limit, 100)

    try:
        query = db.query(Alarm)
        filters = []
        if alarm_type:
            filters.append(Alarm.alarm_type == alarm_type)
        if camera_id:
            filters.append(Alarm.camera_id == camera_id)
        if start_date:
            filters.append(Alarm.alarm_time >= start_date)
        if end_date:
            filters.append(Alarm.alarm_time <= end_date)
        if filters:
            query = query.filter(and_(*filters))

        alarms = query.order_by(desc(Alarm.alarm_time)).limit(limit).all()
        print(f"[Excel Export] Starting export for {len(alarms)} alarms")

        # Get MinIO storage for presigned URLs
        storage = get_minio_storage()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bukti Foto"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ["No", "Foto", "Waktu", "Kamera", "Lokasi", "Tipe Alarm", "Confidence"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[1].height = 25
        img_width, img_height, row_height = 120, 80, 65

        # Try to import Pillow for image embedding
        try:
            from openpyxl.drawing.image import Image as XLImage
            pillow_available = True
        except ImportError:
            pillow_available = False
            print("[Excel Export] Pillow not available, images will show as links")

        # Use a single HTTP client for all requests
        async with httpx.AsyncClient(timeout=5.0) as client:
            for idx, alarm in enumerate(alarms, 1):
                row = idx + 1
                ws.row_dimensions[row].height = row_height
                ws.cell(row=row, column=1, value=idx).border = thin_border
                ws.cell(row=row, column=1).alignment = cell_alignment
                ws.cell(row=row, column=2, value="").border = thin_border
                ws.cell(row=row, column=3, value=alarm.alarm_time.strftime('%d %b %Y\n%H:%M:%S') if alarm.alarm_time else '').border = thin_border
                ws.cell(row=row, column=3).alignment = cell_alignment
                ws.cell(row=row, column=4, value=alarm.camera_name or 'Unknown').border = thin_border
                ws.cell(row=row, column=4).alignment = cell_alignment
                ws.cell(row=row, column=5, value=alarm.location or '-').border = thin_border
                ws.cell(row=row, column=5).alignment = cell_alignment
                ws.cell(row=row, column=6, value=alarm.alarm_type or '').border = thin_border
                ws.cell(row=row, column=6).alignment = cell_alignment
                ws.cell(row=row, column=7, value=f"{round(alarm.confidence * 100)}%" if alarm.confidence else '-').border = thin_border
                ws.cell(row=row, column=7).alignment = cell_alignment

                image_url = None
                is_labeled_image = False
                if alarm.minio_labeled_image_path and storage.is_initialized:
                    image_url = storage.get_presigned_url(settings.minio_bucket_alarm_images, alarm.minio_labeled_image_path)
                    is_labeled_image = True
                elif alarm.minio_image_path and storage.is_initialized:
                    image_url = storage.get_presigned_url(settings.minio_bucket_alarm_images, alarm.minio_image_path)
                elif alarm.image_url:
                    image_url = alarm.image_url

                if image_url and pillow_available:
                    try:
                        img_response = await client.get(image_url)
                        if img_response.status_code == 200:
                            image_content = img_response.content
                            # Add overlay (timestamp + bounding box) for non-labeled images
                            if not is_labeled_image and alarm.alarm_time:
                                image_content = _add_timestamp_overlay(
                                    image_content,
                                    alarm.alarm_time,
                                    alarm.camera_name,
                                    alarm.raw_data,
                                    alarm.alarm_type
                                )
                            img = XLImage(BytesIO(image_content))
                            img.width, img.height = img_width, img_height
                            ws.add_image(img, f"B{row}")
                    except Exception as e:
                        print(f"[Excel] Failed to embed image {idx}: {e}")
                        ws.cell(row=row, column=2, value="(gagal load)").alignment = cell_alignment
                elif image_url and not pillow_available:
                    # Show URL as text if Pillow not available
                    ws.cell(row=row, column=2, value="(lihat link)").alignment = cell_alignment

        column_widths = [5, 18, 15, 20, 25, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        print(f"[Excel Export] Export completed successfully")
        filename = f"bukti_foto_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        print(f"[Excel Export] Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to export: {str(e)}")


@router.get("/{alarm_id}", response_model=AlarmResponse)
def get_alarm(
    alarm_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a specific alarm by ID"""
    alarm = db.query(Alarm).filter(Alarm.id == alarm_id).first()
    if not alarm:
        raise HTTPException(status_code=404, detail="Alarm not found")
    return _add_presigned_urls(alarm)


@router.patch("/{alarm_id}/acknowledge", response_model=AlarmResponse)
def acknowledge_alarm(
    alarm_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Acknowledge an alarm"""
    alarm = db.query(Alarm).filter(Alarm.id == alarm_id).first()
    if not alarm:
        raise HTTPException(status_code=404, detail="Alarm not found")

    alarm.status = "acknowledged"
    alarm.acknowledged_at = datetime.utcnow()
    alarm.acknowledged_by_id = current_user.id
    db.commit()
    db.refresh(alarm)
    return _add_presigned_urls(alarm)


@router.patch("/{alarm_id}/resolve", response_model=AlarmResponse)
def resolve_alarm(
    alarm_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Resolve an alarm"""
    alarm = db.query(Alarm).filter(Alarm.id == alarm_id).first()
    if not alarm:
        raise HTTPException(status_code=404, detail="Alarm not found")

    alarm.status = "resolved"
    alarm.resolved_at = datetime.utcnow()
    alarm.resolved_by_id = current_user.id
    db.commit()
    db.refresh(alarm)
    return _add_presigned_urls(alarm)


@router.delete("/{alarm_id}")
def delete_alarm(
    alarm_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete an alarm"""
    alarm = db.query(Alarm).filter(Alarm.id == alarm_id).first()
    if not alarm:
        raise HTTPException(status_code=404, detail="Alarm not found")

    db.delete(alarm)
    db.commit()
    return {"message": "Alarm deleted"}


@router.post("/bulk-acknowledge")
def bulk_acknowledge(
    alarm_ids: List[UUID],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Acknowledge multiple alarms"""
    now = datetime.utcnow()
    updated = db.query(Alarm).filter(Alarm.id.in_(alarm_ids)).update({
        "status": "acknowledged",
        "acknowledged_at": now,
        "acknowledged_by_id": current_user.id
    }, synchronize_session=False)
    db.commit()
    return {"acknowledged": updated}


@router.post("/bulk-resolve")
def bulk_resolve(
    alarm_ids: List[UUID],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Resolve multiple alarms"""
    now = datetime.utcnow()
    updated = db.query(Alarm).filter(Alarm.id.in_(alarm_ids)).update({
        "status": "resolved",
        "resolved_at": now,
        "resolved_by_id": current_user.id
    }, synchronize_session=False)
    db.commit()
    return {"resolved": updated}


@router.websocket("/ws")
async def alarm_websocket(websocket: WebSocket):
    """WebSocket endpoint for real-time alarm updates"""
    await websocket.accept()
    add_client(websocket)

    try:
        while True:
            # Keep connection alive, receive any client messages
            data = await websocket.receive_text()
            # Client can send ping or commands
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        pass
    finally:
        remove_client(websocket)


# Helper function to save base64 image to MinIO
def _save_base64_image_to_minio(base64_data: str, prefix: str = "alarm") -> Optional[str]:
    """Decode base64 image and save to MinIO. Returns the object path or None."""
    if not base64_data:
        print(f"[MinIO] No base64 data for {prefix}")
        return None

    # Check data length for debugging
    data_len = len(base64_data) if base64_data else 0
    print(f"[MinIO] Received {prefix} base64 data: {data_len} chars")

    storage = get_minio_storage()
    if not storage.is_initialized:
        print(f"[MinIO] Storage not initialized! Check MINIO_ENABLED and connection")
        return None

    try:
        # Decode base64 to bytes
        image_bytes = base64.b64decode(base64_data)
        print(f"[MinIO] Decoded {prefix}: {len(image_bytes)} bytes")

        # Generate unique object name
        object_name = storage.generate_object_name(prefix, "jpg")

        # Upload to MinIO
        result = storage.upload_bytes(
            settings.minio_bucket_alarm_images,
            object_name,
            image_bytes,
            "image/jpeg"
        )

        if result:
            print(f"[MinIO] Saved {prefix}: {object_name}")
            return object_name
        else:
            print(f"[MinIO] Upload returned None for {prefix}")
    except Exception as e:
        print(f"[MinIO] Failed to save {prefix}: {e}")

    return None


# Internal function to save alarm from BM-APP
async def save_alarm_from_bmapp(alarm_data: dict, db: Session):
    """Save an alarm received from BM-APP to database"""
    from dateutil.parser import parse as parse_datetime
    from app.services.telegram import telegram

    alarm_time = alarm_data.get("alarm_time")
    if isinstance(alarm_time, str):
        try:
            alarm_time = parse_datetime(alarm_time)
        except:
            alarm_time = datetime.utcnow()

    # Save base64 images to MinIO if available
    minio_image_path = None
    minio_labeled_image_path = None

    # Debug: Check what image data we received
    image_data_base64 = alarm_data.get("image_data_base64")
    labeled_image_data_base64 = alarm_data.get("labeled_image_data_base64")
    print(f"[Alarm] Image data check: raw={len(image_data_base64) if image_data_base64 else 0} chars, labeled={len(labeled_image_data_base64) if labeled_image_data_base64 else 0} chars")

    # Save raw image
    if image_data_base64:
        minio_image_path = _save_base64_image_to_minio(image_data_base64, "alarm_raw")
    else:
        print("[Alarm] No raw image base64 data received from BM-APP")

    # Save labeled image (with detection boxes) - prioritize this!
    if labeled_image_data_base64:
        minio_labeled_image_path = _save_base64_image_to_minio(labeled_image_data_base64, "alarm_labeled")
    else:
        print("[Alarm] No labeled image base64 data received from BM-APP")

    # Parse aibox_id if provided (comes as string from parsed alarm)
    aibox_id = alarm_data.get("aibox_id")
    if aibox_id and isinstance(aibox_id, str):
        try:
            from uuid import UUID
            aibox_id = UUID(aibox_id)
        except ValueError:
            aibox_id = None

    alarm = Alarm(
        bmapp_id=alarm_data.get("bmapp_id"),
        alarm_type=alarm_data.get("alarm_type", "Unknown"),
        alarm_name=alarm_data.get("alarm_name", "Detection Alert"),
        camera_id=alarm_data.get("camera_id"),
        camera_name=alarm_data.get("camera_name"),
        location=alarm_data.get("location"),
        confidence=float(alarm_data.get("confidence", 0) or 0),
        image_url=alarm_data.get("image_url"),
        video_url=alarm_data.get("video_url"),
        media_url=alarm_data.get("media_url"),  # RTSP URL
        description=alarm_data.get("description"),
        raw_data=alarm_data.get("raw_data"),
        alarm_time=alarm_time,
        status="new",
        # MinIO paths
        minio_image_path=minio_image_path,
        minio_labeled_image_path=minio_labeled_image_path,
        minio_synced_at=datetime.utcnow() if (minio_image_path or minio_labeled_image_path) else None,
        # AI Box info
        aibox_id=aibox_id,
        aibox_name=alarm_data.get("aibox_name"),
    )
    db.add(alarm)
    db.commit()

    # Send Telegram notification (async, non-blocking)
    try:
        # Get image URL for Telegram (prefer labeled image with detection boxes)
        storage = get_minio_storage()
        image_url_for_telegram = None
        if minio_labeled_image_path and storage.is_initialized:
            image_url_for_telegram = storage.get_presigned_url(
                settings.minio_bucket_alarm_images,
                minio_labeled_image_path
            )
        elif minio_image_path and storage.is_initialized:
            image_url_for_telegram = storage.get_presigned_url(
                settings.minio_bucket_alarm_images,
                minio_image_path
            )
        elif alarm_data.get("image_url"):
            image_url_for_telegram = alarm_data.get("image_url")

        await telegram.send_alarm_notification(
            alarm_type=alarm.alarm_type,
            alarm_name=alarm.alarm_name,
            camera_name=alarm.camera_name,
            location=alarm.location,
            alarm_time=alarm_time if isinstance(alarm_time, datetime) else datetime.utcnow(),
            confidence=alarm.confidence,
            image_url=image_url_for_telegram,
            aibox_name=alarm.aibox_name
        )
    except Exception as e:
        print(f"[Alarm] Failed to send Telegram notification: {e}")

    return alarm


@router.post("/test", response_model=AlarmResponse)
async def create_test_alarm(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a test alarm to verify the system is working"""
    import uuid

    test_alarm_data = {
        "bmapp_id": str(uuid.uuid4()),
        "alarm_type": "NoHelmet",
        "alarm_name": "No Helmet Detected",
        "camera_id": "cam_01",
        "camera_name": "Front Gate Camera",
        "location": "Main Entrance - Zone A",
        "confidence": 0.92,
        "image_url": "",
        "video_url": "",
        "description": "Worker detected without safety helmet in restricted area",
        "alarm_time": datetime.utcnow().isoformat()
    }

    alarm = await save_alarm_from_bmapp(test_alarm_data, db)
    return alarm


@router.post("/simulate-bmapp")
async def simulate_bmapp_alarm(
    raw_data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Simulate receiving an alarm from BM-APP to test parsing.

    Send BM-APP format:
    {
        "AlarmId": "uuid",
        "TaskSession": "task_001",
        "TaskDesc": "Helmet Detection",
        "Time": "2024-01-15 10:30:00",
        "Media": {
            "MediaName": "1",
            "MediaDesc": "Front Gate"
        },
        "Result": {
            "Type": "NoHelmet",
            "Description": "No helmet detected"
        }
    }
    """
    # Use the same parsing logic as the listener
    listener = BmAppAlarmListener()
    parsed = listener._parse_alarm(raw_data)

    # Save to database
    alarm = await save_alarm_from_bmapp(parsed, db)

    return {
        "message": "Alarm simulated successfully",
        "raw_input": raw_data,
        "parsed_output": parsed,
        "saved_alarm_id": str(alarm.id)
    }


# ============================================================================
# BM-APP HTTP REPORTING ENDPOINT (NO AUTH - Called directly by BM-APP device)
# ============================================================================

@router.post("/receive")
async def receive_bmapp_alarm(
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Receive alarm from BM-APP via HTTP POST.

    This endpoint is called by BM-APP when MetadataUrl is configured.
    NO AUTHENTICATION required because BM-APP device cannot login.

    Configure in BM-APP task:
        MetadataUrl: "http://YOUR_BACKEND_IP:PORT/api/alarms/receive"

    Expected BM-APP format:
    {
        "BoardId": "RJ-BOX-XXX",
        "AlarmId": "uuid",
        "TaskSession": "task_001",
        "TaskDesc": "Helmet Detection",
        "Time": "2025-01-15 10:30:00",
        "TimeStamp": 1699426698084625,
        "Media": {
            "MediaName": "1",
            "MediaUrl": "rtsp://...",
            "MediaDesc": "H8C-1"
        },
        "Result": {
            "Type": "NoHelmet",
            "Description": "No helmet detected",
            "Properties": [
                {"property": "confidence", "value": 0.683594}
            ]
        },
        "ImageData": "base64...",
        "ImageDataLabeled": "base64..."
    }
    """
    try:
        raw_data = await request.json()

        # Log raw alarm for debugging
        print(f"[BM-APP HTTP] Received alarm: {json.dumps(raw_data, indent=2, default=str)[:1000]}...")

        # Parse using BmAppAlarmListener
        listener = BmAppAlarmListener()
        parsed = listener._parse_alarm(raw_data)

        print(f"[BM-APP HTTP] Parsed: type={parsed.get('alarm_type')}, camera={parsed.get('camera_name')}, conf={parsed.get('confidence')}")

        # Save to database
        alarm = await save_alarm_from_bmapp(parsed, db)

        # Generate MinIO presigned URLs for real-time notification
        storage = get_minio_storage()
        minio_labeled_image_url = None
        minio_image_url = None

        if storage.is_initialized:
            if alarm.minio_labeled_image_path:
                minio_labeled_image_url = storage.get_presigned_url(
                    settings.minio_bucket_alarm_images,
                    alarm.minio_labeled_image_path
                )
            if alarm.minio_image_path:
                minio_image_url = storage.get_presigned_url(
                    settings.minio_bucket_alarm_images,
                    alarm.minio_image_path
                )

        # Broadcast to WebSocket clients for real-time updates
        await broadcast_alarm({
            "id": str(alarm.id),
            "bmapp_id": alarm.bmapp_id,
            "alarm_type": alarm.alarm_type,
            "alarm_name": alarm.alarm_name,
            "camera_id": alarm.camera_id,
            "camera_name": alarm.camera_name,
            "location": alarm.location,
            "confidence": alarm.confidence,
            "image_url": alarm.image_url,
            "video_url": alarm.video_url,
            "media_url": alarm.media_url,  # RTSP URL for video source
            "description": alarm.description,
            "alarm_time": alarm.alarm_time.isoformat() if alarm.alarm_time else None,
            "status": alarm.status,
            # MinIO presigned URLs for images
            "minio_image_url": minio_image_url,
            "minio_labeled_image_url": minio_labeled_image_url,
            "minio_image_path": alarm.minio_image_path,
            "minio_labeled_image_path": alarm.minio_labeled_image_path,
        })

        print(f"[BM-APP HTTP] Alarm saved with ID: {alarm.id}")

        # Return response in BM-APP expected format
        return {
            "Result": {
                "Code": 0,
                "Desc": "Alarm received successfully"
            },
            "AlarmId": str(alarm.id)
        }

    except json.JSONDecodeError as e:
        print(f"[BM-APP HTTP] JSON parse error: {e}")
        return {
            "Result": {
                "Code": 1,
                "Desc": f"Invalid JSON: {str(e)}"
            }
        }
    except Exception as e:
        print(f"[BM-APP HTTP] Error processing alarm: {e}")
        return {
            "Result": {
                "Code": 2,
                "Desc": f"Error: {str(e)}"
            }
        }


# ============================================================================
# DOWNLOAD & EXPORT ENDPOINTS
# ============================================================================

from fastapi.responses import StreamingResponse, Response
from io import BytesIO
import zipfile
import httpx


def _add_timestamp_overlay(
    image_bytes: bytes,
    timestamp: datetime,
    camera_name: str = None,
    raw_data: str = None,
    alarm_type: str = None
) -> bytes:
    """Add timestamp overlay and bounding box to an image.

    Args:
        image_bytes: Original image as bytes
        timestamp: The alarm/capture time to display
        camera_name: Optional camera name to display
        raw_data: Optional raw JSON data from BM-APP containing RelativeBox
        alarm_type: Optional alarm type for label

    Returns:
        Modified image with timestamp and bounding box overlay as bytes
    """
    try:
        from PIL import Image, ImageDraw, ImageFont

        # Open image from bytes
        img = Image.open(BytesIO(image_bytes))

        # Convert to RGB if necessary (for PNG with transparency)
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        img_width, img_height = img.size

        # Calculate font size based on image width
        font_size = max(16, int(img_width * 0.025))
        small_font_size = max(12, int(img_width * 0.018))

        # Try to load fonts
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size)
            small_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", small_font_size)
        except:
            try:
                font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", font_size)
                small_font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", small_font_size)
            except:
                try:
                    font = ImageFont.truetype("arial.ttf", font_size)
                    small_font = ImageFont.truetype("arial.ttf", small_font_size)
                except:
                    font = ImageFont.load_default()
                    small_font = font

        # Create overlay for semi-transparent elements
        overlay = Image.new('RGBA', img.size, (0, 0, 0, 0))
        overlay_draw = ImageDraw.Draw(overlay)

        # ===== DRAW BOUNDING BOX FROM RAW_DATA =====
        if raw_data:
            try:
                data = json.loads(raw_data) if isinstance(raw_data, str) else raw_data
                result = data.get("Result", {})
                relative_box = result.get("RelativeBox")

                if relative_box and len(relative_box) == 4:
                    # RelativeBox format: [x, y, width, height] in relative coords (0-1)
                    rel_x, rel_y, rel_w, rel_h = relative_box

                    # Convert to absolute pixel coordinates
                    box_x = int(rel_x * img_width)
                    box_y = int(rel_y * img_height)
                    box_w = int(rel_w * img_width)
                    box_h = int(rel_h * img_height)

                    # Draw bounding box (red color, 3px width)
                    box_color = (255, 0, 0, 255)  # Red
                    line_width = max(2, int(img_width * 0.004))

                    # Draw rectangle outline
                    for i in range(line_width):
                        overlay_draw.rectangle(
                            [box_x - i, box_y - i, box_x + box_w + i, box_y + box_h + i],
                            outline=box_color
                        )

                    # Draw label background above the box
                    label_text = alarm_type or result.get("Type", "Detection")
                    label_bbox = overlay_draw.textbbox((0, 0), label_text, font=small_font)
                    label_w = label_bbox[2] - label_bbox[0]
                    label_h = label_bbox[3] - label_bbox[1]

                    label_x = box_x
                    label_y = box_y - label_h - 8
                    if label_y < 0:
                        label_y = box_y + box_h + 4  # Put below if no space above

                    # Draw label background
                    overlay_draw.rectangle(
                        [label_x - 2, label_y - 2, label_x + label_w + 6, label_y + label_h + 4],
                        fill=(255, 0, 0, 200)
                    )

                    # Draw label text (will be drawn after compositing)

            except Exception as e:
                print(f"[Download] Failed to parse bounding box: {e}")

        # ===== DRAW TIMESTAMP =====
        timestamp_text = timestamp.strftime('%Y-%m-%d %H:%M:%S')
        if camera_name:
            timestamp_text = f"{camera_name} | {timestamp_text}"

        # Get text bounding box
        draw_temp = ImageDraw.Draw(img)
        bbox = draw_temp.textbbox((0, 0), timestamp_text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]

        # Position: bottom-left with padding
        padding = 10
        x = padding
        y = img_height - text_height - padding - 5

        # Draw semi-transparent background for timestamp
        bg_rect = [x - 5, y - 3, x + text_width + 5, y + text_height + 3]
        overlay_draw.rectangle(bg_rect, fill=(0, 0, 0, 180))

        # Composite overlay onto image
        img = img.convert('RGBA')
        img = Image.alpha_composite(img, overlay)
        img = img.convert('RGB')

        # Draw text elements on final image
        draw = ImageDraw.Draw(img)

        # Draw timestamp text
        draw.text((x, y), timestamp_text, font=font, fill=(255, 255, 255))

        # Draw bounding box label text (if we have box data)
        if raw_data:
            try:
                data = json.loads(raw_data) if isinstance(raw_data, str) else raw_data
                result = data.get("Result", {})
                relative_box = result.get("RelativeBox")

                if relative_box and len(relative_box) == 4:
                    rel_x, rel_y, rel_w, rel_h = relative_box
                    box_x = int(rel_x * img_width)
                    box_y = int(rel_y * img_height)
                    box_h = int(rel_h * img_height)

                    label_text = alarm_type or result.get("Type", "Detection")
                    label_bbox = draw.textbbox((0, 0), label_text, font=small_font)
                    label_h = label_bbox[3] - label_bbox[1]

                    label_x = box_x
                    label_y = box_y - label_h - 8
                    if label_y < 0:
                        label_y = box_y + box_h + 4

                    draw.text((label_x + 2, label_y), label_text, font=small_font, fill=(255, 255, 255))
            except:
                pass

        # Save to bytes
        output = BytesIO()
        img.save(output, format='JPEG', quality=95)
        output.seek(0)

        return output.getvalue()

    except ImportError:
        print("[Download] Pillow not available, returning original image")
        return image_bytes
    except Exception as e:
        print(f"[Download] Failed to add timestamp overlay: {e}")
        return image_bytes


def _get_alarm_image_url(alarm: Alarm, db: Session = None) -> tuple[Optional[str], bool]:
    """Get the best available image URL for an alarm.

    Returns:
        Tuple of (image_url, is_labeled) where is_labeled indicates if the image
        already has bounding boxes drawn by the AI.
    """
    storage = get_minio_storage()

    # Priority 1: MinIO labeled image (with detection boxes)
    if storage.is_initialized and alarm.minio_labeled_image_path:
        return storage.get_presigned_url(
            settings.minio_bucket_alarm_images,
            alarm.minio_labeled_image_path
        ), True  # Already has boxes

    # Priority 2: MinIO raw image
    if storage.is_initialized and alarm.minio_image_path:
        return storage.get_presigned_url(
            settings.minio_bucket_alarm_images,
            alarm.minio_image_path
        ), False  # Raw image, needs boxes

    # Priority 3: BM-APP image_url
    if alarm.image_url:
        if alarm.image_url.startswith(('http://', 'https://')):
            return alarm.image_url, False
        # Construct BM-APP URL for relative paths - use AI Box's api_url if available
        bmapp_base = None
        if alarm.aibox_id and db:
            aibox = db.query(AIBox).filter(AIBox.id == alarm.aibox_id).first()
            if aibox:
                bmapp_base = aibox.api_url.rsplit('/api', 1)[0] if aibox.api_url else None
        if not bmapp_base:
            # Fallback to config (deprecated)
            bmapp_base = settings.bmapp_api_url.rsplit('/api', 1)[0]
        return f"{bmapp_base}/{alarm.image_url.lstrip('/')}", False

    return None, False


@router.get("/{alarm_id}/download-image")
async def download_alarm_image(
    alarm_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Download the alarm image with timestamp overlay."""
    alarm = db.query(Alarm).filter(Alarm.id == alarm_id).first()
    if not alarm:
        raise HTTPException(status_code=404, detail="Alarm not found")

    image_url, is_labeled = _get_alarm_image_url(alarm, db)
    if not image_url:
        raise HTTPException(status_code=404, detail="No image available for this alarm")

    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(image_url)
            response.raise_for_status()

            # Add timestamp and bounding box overlay to image
            # Only draw bounding box if using raw image (labeled already has boxes)
            image_with_overlay = _add_timestamp_overlay(
                response.content,
                alarm.alarm_time,
                alarm.camera_name,
                alarm.raw_data if not is_labeled else None,  # Only pass raw_data for raw images
                alarm.alarm_type if not is_labeled else None
            )

            # Generate filename
            filename = f"alarm_{alarm.alarm_type}_{alarm.alarm_time.strftime('%Y%m%d_%H%M%S')}.jpg"

            return Response(
                content=image_with_overlay,
                media_type="image/jpeg",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to download image: {str(e)}")


@router.post("/bulk-download")
async def bulk_download_images(
    alarm_ids: List[UUID],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Download multiple alarm images as a ZIP file with timestamp overlays."""
    if not alarm_ids:
        raise HTTPException(status_code=400, detail="No alarm IDs provided")

    if len(alarm_ids) > 100:
        raise HTTPException(status_code=400, detail="Maximum 100 images per download")

    alarms = db.query(Alarm).filter(Alarm.id.in_(alarm_ids)).all()
    if not alarms:
        raise HTTPException(status_code=404, detail="No alarms found")

    # Create ZIP in memory
    zip_buffer = BytesIO()

    async with httpx.AsyncClient(timeout=30.0) as client:
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for alarm in alarms:
                image_url, is_labeled = _get_alarm_image_url(alarm, db)
                if not image_url:
                    continue

                try:
                    response = await client.get(image_url)
                    response.raise_for_status()

                    # Add timestamp and bounding box overlay to each image
                    # Only draw bounding box if using raw image (labeled already has boxes)
                    image_with_overlay = _add_timestamp_overlay(
                        response.content,
                        alarm.alarm_time,
                        alarm.camera_name,
                        alarm.raw_data if not is_labeled else None,
                        alarm.alarm_type if not is_labeled else None
                    )

                    filename = f"{alarm.alarm_type}_{alarm.alarm_time.strftime('%Y%m%d_%H%M%S')}_{str(alarm.id)[:8]}.jpg"
                    zf.writestr(filename, image_with_overlay)
                except Exception as e:
                    print(f"Failed to download image for alarm {alarm.id}: {e}")
                    continue

    zip_buffer.seek(0)

    # Generate ZIP filename
    zip_filename = f"alarms_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.zip"

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{zip_filename}"'
        }
    )


